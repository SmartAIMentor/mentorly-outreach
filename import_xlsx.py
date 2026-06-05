#!/usr/bin/env python3
"""从 运营事项.xlsx 导出：运营任务 CSV + TikTok 达人邮件清单。"""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent.parent
XLSX = ROOT / "运营事项.xlsx"
OUT_CREATORS = Path(__file__).resolve().parent / "creators_tiktok.csv"
OUT_TASKS = ROOT / "docs" / "运营事项_更新.csv"


def import_echotik_creators() -> int:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["echotik达人邮箱"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    fieldnames = [
        "platform",
        "handle_or_url",
        "email",
        "niche",
        "name",
        "country",
        "followers",
        "engagement_rate_pct",
        "fastmoss_url",
        "notes",
    ]
    count = 0
    with OUT_CREATORS.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for row in rows:
            if not row or not row[1] or not row[12]:
                continue
            email = str(row[1]).strip()
            if "@" not in email:
                continue
            tiktok_url = str(row[12]).strip()
            if "tiktok.com" not in tiktok_url.lower():
                continue
            niche = str(row[3] or "").strip().replace(",", "/")[:80]
            followers = row[13] if isinstance(row[13], (int, float)) else ""
            eng = row[15] if isinstance(row[15], (int, float)) else ""
            eng_pct = round(float(eng) * 100, 2) if eng != "" and eng != "-" else ""
            w.writerow(
                {
                    "platform": "tiktok",
                    "handle_or_url": tiktok_url,
                    "email": email,
                    "niche": niche or "creator",
                    "name": str(row[0] or row[2] or "").strip(),
                    "country": str(row[5] or "").strip(),
                    "followers": int(followers) if followers != "" else "",
                    "engagement_rate_pct": eng_pct,
                    "fastmoss_url": str(row[11] or "").strip(),
                    "notes": f"handle=@{row[2]}",
                }
            )
            count += 1
    return count


def import_ops_tasks() -> int:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["运营事项"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    fieldnames = [
        "序号",
        "任务名称",
        "渠道类型",
        "优先级",
        "自动化程度",
        "负责人",
        "频率",
        "钩子/卖点",
        "执行步骤",
        "KPI",
        "工具",
        "状态",
        "备注",
    ]
    count = 0
    OUT_TASKS.parent.mkdir(parents=True, exist_ok=True)
    with OUT_TASKS.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for row in rows:
            if not row or not row[1]:
                continue
            w.writerow(
                {
                    "序号": row[0],
                    "任务名称": row[1],
                    "渠道类型": row[2],
                    "优先级": row[3],
                    "自动化程度": row[4],
                    "负责人": row[5],
                    "频率": row[6],
                    "钩子/卖点": row[7],
                    "执行步骤": row[8],
                    "KPI": row[9],
                    "工具": row[10],
                    "状态": row[11],
                    "备注": row[14] if len(row) > 14 else "",
                }
            )
            count += 1
    return count


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"找不到: {XLSX}")
    n1 = import_echotik_creators()
    n2 = import_ops_tasks()
    print(f"✅ TikTok 达人+邮箱: {OUT_CREATORS} ({n1} 人)")
    print(f"✅ 运营任务表: {OUT_TASKS} ({n2} 条)")


if __name__ == "__main__":
    main()
