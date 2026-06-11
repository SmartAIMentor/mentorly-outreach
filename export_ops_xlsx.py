#!/usr/bin/env python3
"""将 docs/运营事项_更新.csv 写回 运营事项.xlsx 的「运营事项」sheet（保留 echotik 达人邮箱）。"""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent
_MENTORAIXS = ROOT.parent.parent


def _paths() -> tuple[Path, Path]:
    """Mentoraixs: scripts/outreach → repo root；mentorly-outreach: 脚本在仓库根目录。"""
    csv_mentoraixs = _MENTORAIXS / "docs" / "运营事项_更新.csv"
    if csv_mentoraixs.exists():
        return csv_mentoraixs, _MENTORAIXS / "运营事项.xlsx"
    return ROOT / "运营事项_更新.csv", _MENTORAIXS / "运营事项.xlsx"


CSV_PATH, XLSX_PATH = _paths()
SHEET_OPS = "运营事项"
SHEET_CREATORS = "echotik达人邮箱"


def load_ops_rows() -> tuple[list[str], list[list]]:
    with CSV_PATH.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def export_xlsx() -> None:
    headers, data = load_ops_rows()
    if not headers:
        raise SystemExit(f"空 CSV: {CSV_PATH}")

    if XLSX_PATH.exists():
        wb = openpyxl.load_workbook(XLSX_PATH)
        if SHEET_OPS in wb.sheetnames:
            del wb[SHEET_OPS]
        ws = wb.create_sheet(SHEET_OPS, 0)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_OPS

    ws.append(headers)
    for row in data:
        ws.append(row)

    wb.save(XLSX_PATH)
    print(f"✅ 已更新 {XLSX_PATH} → sheet「{SHEET_OPS}」({len(data)} 条)")


if __name__ == "__main__":
    export_xlsx()
